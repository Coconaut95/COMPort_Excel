import sys
import os
import xlwings as xw
from openpyxl import load_workbook
from datetime import datetime
from dic_to_list import list_e_fase1, list_e_fase2, list_e_fase3,  list_phi_f1

# absoluta ubicación de este archivo .py
abs_path = os.path.dirname(__file__)

# Path donde se van a guardar las planillas cargadas con datos
relative_path_file = "/resultados"
full_path_file = abs_path + relative_path_file
if not os.path.exists(full_path_file):
    os.mkdir(full_path_file)

# Path Planilla SCVA
relative_path_template_scva = "\Planillas\Planilla_SCVA.xlsx"
full_path_template = abs_path + relative_path_template_scva

# Path Planilla p/normativa directos
relative_path_template_norma = "\Planillas\Planilla_Norma.xls"
full_path_template_norma = abs_path + relative_path_template_norma

# Path Planilla p/normativa indirectos
relative_path_template_norma_ind = "\Planillas\Planilla_Norma_redux.xlsx"
full_path_template_norma_ind = abs_path + relative_path_template_norma_ind

def write_excel_scva(num_serie):
    name_file = full_path_file + '/' + 'SCVA_' + str(num_serie) + '.xlsx'

    if list_phi_f1[0] == '0': # Para identificar si es un archivo de activa.
        wb = load_workbook(filename=full_path_template)
        ws = wb.worksheets[0]
        ws['F4'] = datetime.now().date()
        ws['E14'] = num_serie
        ws['F14'] = list_e_fase1[0]
        ws['G14'] = list_e_fase1[1]
        ws['H14'] = list_e_fase1[2]
        ws['I14'] = list_e_fase1[3]
        ws['J14'] = list_e_fase1[4]
        ws['K14'] = list_e_fase1[5]                # F14, G14 y H14 carga alta fase 1. I14, J14 y K14 carga baja fase 1.

        if list_e_fase2:                                                       # Si pasa el If, es un medidor trifásico.
            ws['E56'] = num_serie
            ws['F56'] = list_e_fase2[0]
            ws['G56'] = list_e_fase2[1]
            ws['H56'] = list_e_fase2[2]
            ws['I56'] = list_e_fase2[3]
            ws['J56'] = list_e_fase2[4]
            ws['K56'] = list_e_fase2[5]

            ws['E98'] = num_serie
            ws['F98'] = list_e_fase3[0]
            ws['G98'] = list_e_fase3[1]
            ws['H98'] = list_e_fase3[2]
            ws['I98'] = list_e_fase3[3]
            ws['J98'] = list_e_fase3[4]
            ws['K98'] = list_e_fase3[5]
    else:                                   # por defecto escribe en la zona de reactiva. No verifica un PF especifico.
        try:
            wb = load_workbook(filename= name_file)
            ws = wb.worksheets[0]
            ws['L14'] = list_e_fase1[0]
            ws['M14'] = list_e_fase1[1]
            ws['N14'] = list_e_fase1[2]
            ws['O14'] = list_e_fase1[3]
            ws['P14'] = list_e_fase1[4]
            ws['Q14'] = list_e_fase1[5]

            ws['L56'] = list_e_fase2[0]
            ws['M56'] = list_e_fase2[1]
            ws['N56'] = list_e_fase2[2]
            ws['O56'] = list_e_fase2[3]
            ws['P56'] = list_e_fase2[4]
            ws['Q56'] = list_e_fase2[5]

            ws['L98'] = list_e_fase3[0]
            ws['M98'] = list_e_fase3[1]
            ws['N98'] = list_e_fase3[2]
            ws['O98'] = list_e_fase3[3]
            ws['P98'] = list_e_fase3[4]
            ws['Q98'] = list_e_fase3[5]
        except FileNotFoundError:
            print("Por favor, ingrese primero archivo de energia activa")
            sys.exit(1)

    wb.save(filename = name_file)

def write_excel_norma(data, num_serie, tipo):

    name_file = full_path_file + '/' + 'Planilla_Norma_' + str(num_serie) + '.xlsx'

    # Tipo de medidores
    tipos_dir = ['DMPA', 'DMGA', 'DMWA', 'DTPA', 'DTGA', 'DTWA']
    tipos_ind = ['DIPA', 'DIGA', 'DIWA']

    if tipo in tipos_ind:
        if data[0]['phi'] == '0':  # Si es de activa, entra al if.
            wb = xw.Book(full_path_template_norma_ind)
            sheet = wb.sheets['Hoja1']
            #Datos
            sheet.range('J6:N6').value = datetime.now().date()
            sheet.range('J9:K9').value = tipo
            sheet.range('M9:N9').value = num_serie
            #Valores de errores p/ cargas activas.
            sheet.range('I27:Q27').value = data[0]['error']
            sheet.range('I38:Q38').value = data[1]['error']
            sheet.range('I33:Q33').value = data[2]['error']
            sheet.range('I16:K16').value = data[3]['error']
            sheet.range('I21:K21').value = data[4]['error']
            sheet.range('L16:N16').value = data[5]['error']
            sheet.range('L21:N21').value = data[6]['error']
            sheet.range('O16:Q16').value = data[7]['error']
            sheet.range('O21:Q21').value = data[8]['error']
            sheet.range('I28:Q28').value = data[9]['error']
            sheet.range('I39:Q39').value = data[10]['error']
            sheet.range('I34:Q34').value = data[11]['error']
            sheet.range('I17:K17').value = data[12]['error']
            sheet.range('I22:K22').value = data[13]['error']
            sheet.range('L17:N17').value = data[14]['error']
            sheet.range('L22:N22').value = data[15]['error']
            sheet.range('O17:Q17').value = data[16]['error']
            sheet.range('O22:Q22').value = data[17]['error']
            sheet.range('I29:Q29').value = data[18]['error']
            sheet.range('I40:Q40').value = data[19]['error']
            sheet.range('I35:Q35').value = data[20]['error']
            sheet.range('I18:K18').value = data[21]['error']
            sheet.range('I23:K23').value = data[22]['error']
            sheet.range('L18:N18').value = data[23]['error']
            sheet.range('L23:N23').value = data[24]['error']
            sheet.range('O18:Q18').value = data[25]['error']
            sheet.range('O23:Q23').value = data[26]['error']
            sheet.range('I30:Q30').value = data[27]['error']
            sheet.range('I41:Q41').value = data[28]['error']
            sheet.range('I36:Q36').value = data[29]['error']
            sheet.range('I19:K19').value = data[30]['error']
            sheet.range('I24:K24').value = data[31]['error']
            sheet.range('L19:N19').value = data[32]['error']
            sheet.range('L24:N24').value = data[33]['error']
            sheet.range('O19:Q19').value = data[34]['error']
            sheet.range('O24:Q24').value = data[35]['error']
            sheet.range('I31:Q31').value = data[36]['error']
            sheet.range('I42:Q42').value = data[37]['error']
            sheet.range('I37:Q37').value = data[38]['error']
            sheet.range('I20:K20').value = data[39]['error']
            sheet.range('L20:N20').value = data[40]['error']
            sheet.range('O20:Q20').value = data[41]['error']
            sheet.range('I32:Q32').value = data[42]['error']
        else:
            try:
                wb = xw.Book(name_file)
                #filas de valores de error p/ carga reactiva
                sheet = wb.sheets['Hoja1']
                sheet.range('I57:Q57').value = data[0]['error']
                sheet.range('I67:Q67').value = data[1]['error']
                sheet.range('I63:Q63').value = data[2]['error']
                sheet.range('I46:K46').value = data[3]['error']
                sheet.range('I51:K51').value = data[4]['error']
                sheet.range('L46:N46').value = data[5]['error']
                sheet.range('L51:N51').value = data[6]['error']
                sheet.range('O46:Q46').value = data[7]['error']
                sheet.range('O51:Q51').value = data[8]['error']
                sheet.range('I58:Q58').value = data[9]['error']
                sheet.range('I68:Q68').value = data[10]['error']
                sheet.range('I64:Q64').value = data[11]['error']
                sheet.range('I47:K47').value = data[12]['error']
                sheet.range('I52:K52').value = data[13]['error']
                sheet.range('L47:N47').value = data[14]['error']
                sheet.range('L52:N52').value = data[15]['error']
                sheet.range('O47:Q47').value = data[16]['error']
                sheet.range('O52:Q52').value = data[17]['error']
                sheet.range('I59:Q59').value = data[18]['error']
                sheet.range('I69:Q69').value = data[19]['error']
                sheet.range('I65:Q65').value = data[20]['error']
                sheet.range('I48:K48').value = data[21]['error']
                sheet.range('I53:K53').value = data[22]['error']
                sheet.range('L48:N48').value = data[23]['error']
                sheet.range('L53:N53').value = data[24]['error']
                sheet.range('O48:Q48').value = data[25]['error']
                sheet.range('O53:Q53').value = data[26]['error']
                sheet.range('I60:Q60').value = data[27]['error']
                sheet.range('I70:Q70').value = data[28]['error']
                sheet.range('I66:Q66').value = data[29]['error']
                sheet.range('I49:K49').value = data[30]['error']
                sheet.range('I54:K54').value = data[31]['error']
                sheet.range('L49:N49').value = data[32]['error']
                sheet.range('L54:N54').value = data[33]['error']
                sheet.range('O49:Q49').value = data[34]['error']
                sheet.range('O54:Q54').value = data[35]['error']
                sheet.range('I61:Q61').value = data[36]['error']
                sheet.range('I71:Q71').value = data[37]['error']
                sheet.range('I50:K50').value = data[38]['error']
                sheet.range('L50:N50').value = data[39]['error']
                sheet.range('O50:Q50').value = data[40]['error']
                sheet.range('I62:Q62').value = data[41]['error']
            except FileNotFoundError:
                print('Por favor, ingrese primero archivo de energía activa ')
                sys.exit(1)

        wb.save(name_file)

    elif tipo in tipos_dir:
        if data[0]['phi'] == '0':
            wb = xw.Book(full_path_template_norma)
            sheet = wb.sheets['Hoja1']
            # Datos
            sheet.range('J6:N6').value = datetime.now().date()
            sheet.range('J9:K9').value = tipo
            sheet.range('M9:N9').value = num_serie
            # Valores de errores p/ cargas activas.
            sheet.range('I43:Q43').value = data[0]['error']
            sheet.range('I70:Q70').value = data[1]['error']
            sheet.range('I57:Q57').value = data[2]['error']
            sheet.range('I16:K16').value = data[3]['error']
            sheet.range('I29:K29').value = data[4]['error']
            sheet.range('L16:N16').value = data[5]['error']
            sheet.range('L29:N29').value = data[6]['error']
            sheet.range('O16:Q16').value = data[7]['error']
            sheet.range('O29:Q29').value = data[8]['error']
            sheet.range('I44:Q44').value = data[9]['error']
            sheet.range('I71:Q71').value = data[10]['error']
            sheet.range('I58:Q58').value = data[11]['error']
            sheet.range('I17:K17').value = data[12]['error']
            sheet.range('I30:K30').value = data[13]['error']
            sheet.range('L17:N17').value = data[14]['error']
            sheet.range('L30:N30').value = data[15]['error']
            sheet.range('O17:Q17').value = data[16]['error']
            sheet.range('O30:Q30').value = data[17]['error']
            sheet.range('I45:Q45').value = data[18]['error']
            sheet.range('I72:Q72').value = data[19]['error']
            sheet.range('I59:Q59').value = data[20]['error']
            sheet.range('I18:K18').value = data[21]['error']
            sheet.range('I31:K31').value = data[22]['error']
            sheet.range('L18:N18').value = data[23]['error']
            sheet.range('L31:N31').value = data[24]['error']
            sheet.range('O18:Q18').value = data[25]['error']
            sheet.range('O31:Q31').value = data[26]['error']
            sheet.range('I46:Q46').value = data[27]['error']
            sheet.range('I73:Q73').value = data[28]['error']
            sheet.range('I60:Q60').value = data[29]['error']
            sheet.range('I19:K19').value = data[30]['error']
            sheet.range('I32:K32').value = data[31]['error']
            sheet.range('L19:N19').value = data[32]['error']
            sheet.range('L32:N32').value = data[33]['error']
            sheet.range('O19:Q19').value = data[34]['error']
            sheet.range('O32:Q32').value = data[35]['error']
            sheet.range('I47:Q47').value = data[36]['error']
            sheet.range('I74:Q74').value = data[37]['error']
            sheet.range('I61:Q61').value = data[38]['error']
            sheet.range('I20:K20').value = data[39]['error']
            sheet.range('I33:K33').value = data[40]['error']
            sheet.range('L20:N20').value = data[41]['error']
            sheet.range('L33:N33').value = data[42]['error']
            sheet.range('O20:Q20').value = data[43]['error']
            sheet.range('O33:Q33').value = data[44]['error']
            sheet.range('I48:Q48').value = data[45]['error']
            sheet.range('I75:Q75').value = data[46]['error']
            sheet.range('I62:Q62').value = data[47]['error']
            sheet.range('I21:K21').value = data[48]['error']
            sheet.range('I34:K34').value = data[49]['error']
            sheet.range('L21:N21').value = data[50]['error']
            sheet.range('L34:N34').value = data[51]['error']
            sheet.range('O21:Q21').value = data[52]['error']
            sheet.range('O34:Q34').value = data[53]['error']
            sheet.range('I49:Q49').value = data[54]['error']
            sheet.range('I76:Q76').value = data[55]['error']
            sheet.range('I63:Q63').value = data[56]['error']
            sheet.range('I22:K22').value = data[57]['error']
            sheet.range('I35:K35').value = data[58]['error']
            sheet.range('L22:N22').value = data[59]['error']
            sheet.range('L35:N35').value = data[60]['error']
            sheet.range('O22:Q22').value = data[61]['error']
            sheet.range('O35:Q35').value = data[62]['error']
            sheet.range('I50:Q50').value = data[63]['error']
            sheet.range('I77:Q77').value = data[64]['error']
            sheet.range('I64:Q64').value = data[65]['error']
            sheet.range('I23:K23').value = data[66]['error']
            sheet.range('I36:K36').value = data[67]['error']
            sheet.range('L23:N23').value = data[68]['error']
            sheet.range('L36:N36').value = data[69]['error']
            sheet.range('O23:Q23').value = data[70]['error']
            sheet.range('O36:Q36').value = data[71]['error']
            sheet.range('I51:Q51').value = data[72]['error']
            sheet.range('I78:Q78').value = data[73]['error']
            sheet.range('I65:Q65').value = data[74]['error']
            sheet.range('I24:K24').value = data[75]['error']
            sheet.range('I37:K37').value = data[76]['error']
            sheet.range('L24:N24').value = data[77]['error']
            sheet.range('L37:N37').value = data[78]['error']
            sheet.range('O24:Q24').value = data[79]['error']
            sheet.range('O37:Q37').value = data[80]['error']
            sheet.range('I52:Q52').value = data[81]['error']
            sheet.range('I79:Q79').value = data[82]['error']
            sheet.range('I66:Q66').value = data[83]['error']
            sheet.range('I25:K25').value = data[84]['error']
            sheet.range('I38:K38').value = data[85]['error']
            sheet.range('L25:N25').value = data[86]['error']
            sheet.range('L38:N38').value = data[87]['error']
            sheet.range('O25:Q25').value = data[88]['error']
            sheet.range('O38:Q38').value = data[89]['error']
            sheet.range('I53:Q53').value = data[90]['error']
            sheet.range('I80:Q80').value = data[91]['error']
            sheet.range('I67:Q67').value = data[92]['error']
            sheet.range('I26:K26').value = data[93]['error']
            sheet.range('I39:K39').value = data[94]['error']
            sheet.range('L26:N26').value = data[95]['error']
            sheet.range('L39:N39').value = data[96]['error']
            sheet.range('O26:Q26').value = data[97]['error']
            sheet.range('O39:Q39').value = data[98]['error']
            sheet.range('I54:Q54').value = data[99]['error']
            sheet.range('I81:Q81').value = data[100]['error']
            sheet.range('I68:Q68').value = data[101]['error']
            sheet.range('I27:K27').value = data[102]['error']
            sheet.range('L27:N27').value = data[103]['error']
            sheet.range('O27:Q27').value = data[104]['error']
            sheet.range('I55:Q55').value = data[105]['error']
        else:
            try:
                wb = xw.Book(name_file)
                #filas de valores de error p/ carga reactiva
                sheet = wb.sheets['Hoja1']
                #data.remove(data[56])
                sheet.range('I107:Q107').value = data[0]['error']
                sheet.range('I127:Q127').value = data[1]['error']
                sheet.range('I118:Q118').value = data[2]['error']
                sheet.range('I86:K86').value = data[3]['error']
                sheet.range('I96:K96').value = data[4]['error']
                sheet.range('L86:N86').value = data[5]['error']
                sheet.range('L96:N96').value = data[6]['error']
                sheet.range('O86:Q86').value = data[7]['error']
                sheet.range('O96:Q96').value = data[8]['error']
                sheet.range('I108:Q108').value = data[9]['error']
                sheet.range('I128:Q128').value = data[10]['error']
                sheet.range('I119:Q119').value = data[11]['error']
                sheet.range('I87:K87').value = data[12]['error']
                sheet.range('I97:K97').value = data[13]['error']
                sheet.range('L87:N87').value = data[14]['error']
                sheet.range('L97:N97').value = data[15]['error']
                sheet.range('O87:Q87').value = data[16]['error']
                sheet.range('O97:Q97').value = data[17]['error']
                sheet.range('I109:Q109').value = data[18]['error']
                sheet.range('I129:Q129').value = data[19]['error']
                sheet.range('I120:Q120').value = data[20]['error']
                sheet.range('I88:K88').value = data[21]['error']
                sheet.range('I98:K98').value = data[22]['error']
                sheet.range('L88:N88').value = data[23]['error']
                sheet.range('L98:N98').value = data[24]['error']
                sheet.range('O88:Q88').value = data[25]['error']
                sheet.range('O98:Q98').value = data[26]['error']
                sheet.range('I110:Q110').value = data[27]['error']
                sheet.range('I130:Q130').value = data[28]['error']
                sheet.range('I121:Q121').value = data[29]['error']
                sheet.range('I89:K89').value = data[30]['error']
                sheet.range('I99:K99').value = data[31]['error']
                sheet.range('L89:N89').value = data[32]['error']
                sheet.range('L99:N99').value = data[33]['error']
                sheet.range('O89:Q89').value = data[34]['error']
                sheet.range('O99:Q99').value = data[35]['error']
                sheet.range('I111:Q111').value = data[36]['error']
                sheet.range('I131:Q131').value = data[37]['error']
                sheet.range('I122:Q122').value = data[38]['error']
                sheet.range('I90:K90').value = data[39]['error']
                sheet.range('I100:K100').value = data[40]['error']
                sheet.range('L90:N90').value = data[41]['error']
                sheet.range('L100:N100').value = data[42]['error']
                sheet.range('O90:Q90').value = data[43]['error']
                sheet.range('O100:Q100').value = data[44]['error']
                sheet.range('I112:Q112').value = data[45]['error']
                sheet.range('I132:Q132').value = data[46]['error']
                sheet.range('I123:Q123').value = data[47]['error']
                sheet.range('I91:K91').value = data[48]['error']
                sheet.range('I101:K101').value = data[49]['error']
                sheet.range('L91:N91').value = data[50]['error']
                sheet.range('L101:N101').value = data[51]['error']
                sheet.range('O91:Q91').value = data[52]['error']
                sheet.range('O101:Q101').value = data[53]['error']
                sheet.range('I113:Q113').value = data[54]['error']
                sheet.range('I133:Q133').value = data[55]['error']
                sheet.range('I124:Q124').value = data[57]['error']
                sheet.range('I92:K92').value = data[58]['error']
                sheet.range('I102:K102').value = data[59]['error']
                sheet.range('L92:N92').value = data[60]['error']
                sheet.range('L102:N102').value = data[61]['error']
                sheet.range('O92:Q92').value = data[62]['error']
                sheet.range('O102:Q102').value = data[63]['error']
                sheet.range('I114:Q114').value = data[64]['error']
                sheet.range('I134:Q134').value = data[65]['error']
                sheet.range('I125:Q125').value = data[66]['error']
                sheet.range('I93:K93').value = data[67]['error']
                sheet.range('I103:K103').value = data[68]['error']
                sheet.range('L93:N93').value = data[69]['error']
                sheet.range('L103:N103').value = data[70]['error']
                sheet.range('O93:Q93').value = data[71]['error']
                sheet.range('O103:Q103').value = data[72]['error']
                sheet.range('I115:Q115').value = data[73]['error']
                sheet.range('I135:Q135').value = data[74]['error']
                sheet.range('I94:K94').value = data[75]['error']
                sheet.range('L94:N94').value = data[76]['error']
                sheet.range('O94:Q94').value = data[77]['error']
                sheet.range('I116:Q116').value = data[78]['error']
            except FileNotFoundError:
                print('Por favor, ingrese primero archivo de energía activa ')
                sys.exit(1)

        wb.save(name_file)



